import os

from cli.interfaces.metadata_container_interface import (
    IMetadataContainer, MetadataContainerException)
from cli.model_creator import ModelCreator
from cli.utils import get_file_extension
from openpyxl import load_workbook


class XLSXContainer(IMetadataContainer):

    def validate_origin(self, origin):
        # Find xlsx files
        origins = []
        if not isinstance(origin, list):
            origin = [origin]
        for file in origin:
            if os.path.isfile(file):
                files = [file] if get_file_extension(file) == '.xlsx' else []

            else:
                files = []
            origins.extend(files)
        if len(origins) > 0:
            self.origin = origins
        else:
            raise MetadataContainerException(
                "No .xlsx files found in {0}".format(origin))

    def get_model_creators(self):
        if not self._model_creators:
            self._model_creators = []
            for file in self.origin:
                models = self.process_xls_file(
                    file, self.just_validate, self.ignore_field_errors)
                self._model_creators.extend(models)

        return self._model_creators

    def process_xls_file(self,
                         file_name: str,
                         just_validate: bool,
                         ignore_field_errors: bool):

        wb = load_workbook(filename=file_name)
        models = []
        files_ok = []
        files_error = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            mc = self.process_sheet(sheet, just_validate,
                                    ignore_field_errors, files_ok, files_error)
            if mc and mc.is_ok:
                models.append(mc)

        return models

    def process_sheet(self,
                      sheet,
                      just_validate: bool,
                      ignore_field_errors: bool,
                      files_ok: list,
                      files_error: list) -> ModelCreator:
        # Primeira linha - header
        (promax_ns, promax_model, ms_ns, ms_model) = (
            sheet['A1'].value,
            sheet['B1'].value,
            sheet['C1'].value,
            sheet['D1'].value)

        if '.' in promax_ns and not ms_ns:
            (promax_ns, promax_model, ms_ns, ms_model) = (
                promax_ns.split('.')[0],
                promax_ns.split('.')[1],
                promax_model.split('.')[0],
                promax_model.split('.')[1]
            )

        lines = []

        header = '{0}.{1};{2}.{3}'.format(
            promax_ns,
            promax_model,
            ms_ns,
            ms_model)
        lines.append(header)
        end_process = False
        line = 2
        while not end_process:
            (promax_field, promax_type, ms_field, ms_type, ms_extra) = (
                sheet.cell(column=1, row=line).value,
                sheet.cell(column=2, row=line).value,
                sheet.cell(column=3, row=line).value,
                sheet.cell(column=4, row=line).value,
                sheet.cell(column=5, row=line).value)

            if promax_field and promax_type and ms_field and ms_type:
                field = '{0};{1};{2};{3};{4};'.format(
                    promax_field,
                    promax_type,
                    ms_field,
                    ms_type,
                    ms_extra
                )
                lines.append(field)
            else:
                end_process = True
            line += 1

        mc = ModelCreator(origin=lines,
                          ignore_field_errors=ignore_field_errors,
                          just_validate=just_validate)

        if mc.is_ok:
            files_ok.append(sheet.title)
        else:
            files_error.append(sheet.title)

        return mc
