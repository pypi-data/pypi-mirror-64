#
# Generated FMC REST API sample script
#
import json
import sys
import requests
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Color, Border, Side, PatternFill, Font, GradientFill, Alignment, colors


## Function Connect API FMC ##
def connect(server,username,password):

    global g_server, g_username, g_password, g_domain_uuid, g_headers

    g_server = server
    if (g_server[-1] == '/'):
        g_server = g_server[:-1]

    g_username = username
    g_password = password

    r = None
    g_headers = {'Content-Type': 'application/json'}
    api_auth_path = "/api/fmc_platform/v1/auth/generatetoken"
    auth_url = g_server + api_auth_path
    try:
        # 2 ways of making a REST call are provided:
        # One with "SSL verification turned off" and the other with "SSL verification turned on".
        # The one with "SSL verification turned off" is commented out. If you like to use that then
        # uncomment the line where verify=False and comment the line with =verify='/path/to/ssl_certificate'
        # REST call with SSL verification turned off:
        # r = requests.post(auth_url, headers=headers, auth=requests.auth.HTTPBasicAuth(username,password), verify=False)
        # REST call with SSL verification turned on: Download SSL certificates from your FMC first and provide its path for verification.
        r = requests.post(auth_url, headers=g_headers, auth=requests.auth.HTTPBasicAuth(username,password), verify=False)
        auth_headers = r.headers

        print (r.headers)
        #print ("Domain UUID = " + r.headers["DOMAIN_UUID"])

        auth_token = auth_headers.get('X-auth-access-token', default=None)
        refresh_token = auth_headers.get('X-auth-refresh-token', default=None)

        if auth_token == None and refresh_token == None:
            print("auth_token not found. Exiting...")
            sys.exit()
    except Exception as err:
        print ("Error in generating auth token --> "+str(err))
        sys.exit()

    g_headers['X-auth-access-token'] = auth_token
    g_headers['X-auth-refresh-token'] = refresh_token

    print ("++++++++ headers ++++++++ : "+str(g_headers))

    g_domain_uuid = r.headers["DOMAIN_UUID"]
    print ("++++++++ g_domain_uuid ++++++++ : " + g_domain_uuid)
    print ("-------- Username -------- : " + g_username)
    print ("-------- Password -------- : " + g_password)

def get_url(url):
    
    try:
        r = None
        r = requests.get(url, headers=g_headers, verify=False)
        status_code = r.status_code
        resp = r.text
        
        time.sleep(0.5)

        if (status_code == 200):
            print("GET successful. Response data --> ")
            json_resp = json.loads(resp)
            
            # print(json.dumps(json_resp,sort_keys=True,indent=4, separators=(',', ': ')))

        elif (status_code == 401):
            print("+++++++++ Unauthorized +++++++++")
            r.close()
            connect(g_server, g_username, g_password)
            json_resp = get_url(url)
        else:
            r.raise_for_status()
            print("Error occurred in GET --> "+resp)

    except requests.exceptions.HTTPError as err:
        print ("Error in connection --> "+str(err))

    finally:
        if r : r.close()

    return json_resp

## Function API Get NetworkAddresses ##
def get_networkaddresses():

    offset = 0
    limit = 1000
    index = 2
    paging = 1

    ## Create workbook, sheet in Excel ##
    wb = Workbook()
    ws_NetworkAddresses = wb.active
    ws_NetworkAddresses.title = "NetworkAddresses"

    ## Header NetworkAddresses in Excel ##
    ws_NetworkAddresses["A1"] = "ID_Objects"
    ws_NetworkAddresses["A1"].font = Font(bold=True)
    ws_NetworkAddresses["B1"] = "Name"
    ws_NetworkAddresses["B1"].font = Font(bold=True)
    ws_NetworkAddresses["C1"] = "Type"
    ws_NetworkAddresses["C1"].font = Font(bold=True)
    ws_NetworkAddresses["D1"] = "Value"
    ws_NetworkAddresses["D1"].font = Font(bold=True)
    ws_NetworkAddresses["E1"] = "Description"
    ws_NetworkAddresses["E1"].font = Font(bold=True)
    ws_NetworkAddresses["F1"] = "Overridable"
    ws_NetworkAddresses["F1"].font = Font(bold=True)
    ws_NetworkAddresses["G1"] = "Status"
    ws_NetworkAddresses["G1"].font = Font(bold=True)

    api_path = "/api/fmc_config/v1/domain/" + g_domain_uuid + "/object/networkaddresses?offset=" + str(offset) + "&limit=" +str(limit)    # param
    url = g_server + api_path

    json_resp = get_url(url)

    count = json_resp["paging"]["count"]
    print ("************ Pages total : " + str(json_resp["paging"]["pages"]) + " ************")

    for offset_num in range (offset,count,limit):
        print ("************ paging = " + str(paging) + ", offset = " + str(offset_num) + ", limit = " + str(limit) + " ************")
        api_path = "/api/fmc_config/v1/domain/" + g_domain_uuid + "/object/networkaddresses?offset=" + str(offset_num) + "&limit=" +str(limit)    # param
        url = g_server + api_path

        json_resp = get_url(url)

        for data in json_resp["items"]:
            #print (json.dumps(data['id'],sort_keys=True,indent=4, separators=(',', ': ')))
            print ("---------------------------------------------")

            # Get value in networkaddresses
            url = data["links"]["self"]    # param

            json_resp = get_url(url)

            print(json.dumps(json_resp,sort_keys=True,indent=4, separators=(',', ': ')))

            if not "readOnly" in json_resp["metadata"]:
                # print ("id : " + json_resp["id"])
                # print ("name : " + json_resp["name"])
                # print ("type : " + json_resp["type"])
                # print ("value : " + json_resp["value"])
                # print ("description : " + json_resp["description"])

                ws_NetworkAddresses["A" + str(index)] = json_resp["id"]
                ws_NetworkAddresses["B" + str(index)] = json_resp["name"]
                ws_NetworkAddresses["C" + str(index)] = json_resp["type"]
                ws_NetworkAddresses["D" + str(index)] = json_resp["value"]
                ws_NetworkAddresses["E" + str(index)] = json_resp["description"]
                ws_NetworkAddresses["F" + str(index)] = str(json_resp["overridable"])

                index += 1
            else :
                # print ("id : " + json_resp["id"])
                # print ("name : " + json_resp["name"])
                # print ("type : " + json_resp["type"])
                # print ("value : " + json_resp["value"])
                # print ("description : " + json_resp["description"])

                ws_NetworkAddresses["A" + str(index)] = json_resp["id"]
                ws_NetworkAddresses["B" + str(index)] = json_resp["name"]
                ws_NetworkAddresses["C" + str(index)] = json_resp["type"]
                ws_NetworkAddresses["D" + str(index)] = json_resp["value"]
                ws_NetworkAddresses["E" + str(index)] = json_resp["description"]
                ws_NetworkAddresses["F" + str(index)] = str(json_resp["overridable"])
                ws_NetworkAddresses["G" + str(index)] = json_resp["metadata"]["readOnly"]["reason"]

                index += 1

        paging += 1

    print ("\n")
    print ("NetworkAddresses Total = " + str(count))
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    print("The Desktop path is: " + desktop)
    wb.save(desktop + "/NetworkAddresses.xlsx")

## Function API Get NetworkGroups ##
def get_networkgroups():

    offset = 0
    limit = 1000
    index = 2
    paging = 1

    ## Create workbook, sheet in Excel ##
    wb = Workbook()
    ws_NetworkGroups = wb.active
    ws_NetworkGroups.title = "NetworkGroups"

    ## Header NetworkGroups in Excel ##
    ws_NetworkGroups["A1"] = "ID_Groups"
    ws_NetworkGroups["A1"].font = Font(bold=True)
    ws_NetworkGroups["B1"] = "Name"
    ws_NetworkGroups["B1"].font = Font(bold=True)
    ws_NetworkGroups["C1"] = "Type"
    ws_NetworkGroups["C1"].font = Font(bold=True)
    ws_NetworkGroups["D1"] = "Description"
    ws_NetworkGroups["D1"].font = Font(bold=True)
    ws_NetworkGroups["E1"] = "Overridable"
    ws_NetworkGroups["E1"].font = Font(bold=True)
    ws_NetworkGroups["F1"] = "Status"
    ws_NetworkGroups["F1"].font = Font(bold=True)
    ws_NetworkGroups["G1"] = "Literals_type"
    ws_NetworkGroups["G1"].font = Font(bold=True)
    ws_NetworkGroups["H1"] = "Literals_value"
    ws_NetworkGroups["H1"].font = Font(bold=True)
    ws_NetworkGroups["I1"] = "Objects_ID"
    ws_NetworkGroups["I1"].font = Font(bold=True)
    ws_NetworkGroups["J1"] = "Objects_name"
    ws_NetworkGroups["J1"].font = Font(bold=True)
    ws_NetworkGroups["K1"] = "Objects_type"
    ws_NetworkGroups["K1"].font = Font(bold=True)

    api_path = "/api/fmc_config/v1/domain/" + g_domain_uuid + "/object/networkgroups?offset=" + str(offset) + "&limit=" +str(limit)    # param
    url = g_server + api_path

    json_resp = get_url(url)

    count = json_resp["paging"]["count"]
    print ("************ Pages total : " + str(json_resp["paging"]["pages"]) + " ************")

    for offset_num in range (offset,count,limit):
        print ("************ paging = " + str(paging) + ", offset = " + str(offset_num) + ", limit = " + str(limit) + " ************")
        api_path = "/api/fmc_config/v1/domain/" + g_domain_uuid + "/object/networkgroups?offset=" + str(offset_num) + "&limit=" +str(limit)    # param
        url = g_server + api_path

        json_resp = get_url(url)
     
        for data in json_resp['items']:
            #print (json.dumps(data['id'],sort_keys=True,indent=4, separators=(',', ': ')))
            print ('---------------------------------------------')

            url = data["links"]["self"]    # param

            json_resp = get_url(url)

            print(json.dumps(json_resp,sort_keys=True,indent=4, separators=(',', ': ')))

            if not "readOnly" in json_resp["metadata"]: ## Not readOnly ##

                if "literals" in json_resp and "objects" in json_resp:
                    print ("************** literals and objects **************")
                    index_literals = index
                    index_objects = index
                    for literals in json_resp["literals"]:
                        # print ("literals type : " + literals["type"])
                        # print ("literals value : " + literals["value"])

                        ws_NetworkGroups["A" + str(index_literals)] = json_resp["id"]
                        ws_NetworkGroups["B" + str(index_literals)] = json_resp["name"]
                        ws_NetworkGroups["C" + str(index_literals)] = json_resp["type"]
                        ws_NetworkGroups["D" + str(index_literals)] = json_resp["description"]
                        ws_NetworkGroups["E" + str(index_literals)] = json_resp["overridable"]

                        ws_NetworkGroups["G" + str(index_literals)] = literals["type"]
                        ws_NetworkGroups["H" + str(index_literals)] = literals["value"]

                        index_literals += 1

                    for objects in json_resp["objects"]:
                        # print ("objects id : " + objects["id"])
                        # print ("objects name : " + objects["name"])
                        # print ("objects type : " + objects["type"])

                        ws_NetworkGroups["A" + str(index_objects)] = json_resp["id"]
                        ws_NetworkGroups["B" + str(index_objects)] = json_resp["name"]
                        ws_NetworkGroups["C" + str(index_objects)] = json_resp["type"]
                        ws_NetworkGroups["D" + str(index_objects)] = json_resp["description"]
                        ws_NetworkGroups["E" + str(index_objects)] = json_resp["overridable"]

                        ws_NetworkGroups["I" + str(index_objects)] = objects["id"]
                        ws_NetworkGroups["J" + str(index_objects)] = objects["name"]
                        ws_NetworkGroups["K" + str(index_objects)] = objects["type"]

                        index_objects += 1

                    if index_literals > index_objects:
                        index = index_literals
                    else :
                        index = index_objects

                elif "literals" in json_resp:
                    print ("************** literals **************") 
                    for literals in json_resp["literals"]:
                        # print ("literals type : " + literals["type"])
                        # print ("literals value : " + literals["value"])

                        ws_NetworkGroups["A" + str(index)] = json_resp["id"]
                        ws_NetworkGroups["B" + str(index)] = json_resp["name"]
                        ws_NetworkGroups["C" + str(index)] = json_resp["type"]
                        ws_NetworkGroups["D" + str(index)] = json_resp["description"]
                        ws_NetworkGroups["E" + str(index)] = json_resp["overridable"]

                        ws_NetworkGroups["G" + str(index)] = literals["type"]
                        ws_NetworkGroups["H" + str(index)] = literals["value"]

                        index += 1

                elif "objects" in json_resp:
                    print ("************** objects **************") 
                    for objects in json_resp["objects"]:
                        # print ("objects id : " + objects["id"])
                        # print ("objects name : " + objects["name"])
                        # print ("objects type : " + objects["type"])

                        ws_NetworkGroups["A" + str(index)] = json_resp["id"]
                        ws_NetworkGroups["B" + str(index)] = json_resp["name"]
                        ws_NetworkGroups["C" + str(index)] = json_resp["type"]
                        ws_NetworkGroups["D" + str(index)] = json_resp["description"]
                        ws_NetworkGroups["E" + str(index)] = json_resp["overridable"]

                        ws_NetworkGroups["I" + str(index)] = objects["id"]
                        ws_NetworkGroups["J" + str(index)] = objects["name"]
                        ws_NetworkGroups["K" + str(index)] = objects["type"]

                        index += 1
                else :
                    pass

            else : ## readOnly ##

                if "literals" in json_resp and "objects" in json_resp:
                    print ("************** literals and objects **************")
                    index_literals = index
                    index_objects = index
                    for literals in json_resp["literals"]:
                        # print ("literals type : " + literals["type"])
                        # print ("literals value : " + literals["value"])

                        ws_NetworkGroups["A" + str(index_literals)] = json_resp["id"]
                        ws_NetworkGroups["B" + str(index_literals)] = json_resp["name"]
                        ws_NetworkGroups["C" + str(index_literals)] = json_resp["type"]
                        ws_NetworkGroups["D" + str(index_literals)] = json_resp["description"]
                        ws_NetworkGroups["E" + str(index_literals)] = json_resp["overridable"]
                        ws_NetworkGroups["F" + str(index_literals)] = json_resp["metadata"]["readOnly"]["reason"]

                        ws_NetworkGroups["G" + str(index_literals)] = literals["type"]
                        ws_NetworkGroups["H" + str(index_literals)] = literals["value"]

                        index_literals += 1

                    for objects in json_resp["objects"]:
                        # print ("objects id : " + objects["id"])
                        # print ("objects name : " + objects["name"])
                        # print ("objects type : " + objects["type"])

                        ws_NetworkGroups["A" + str(index_objects)] = json_resp["id"]
                        ws_NetworkGroups["B" + str(index_objects)] = json_resp["name"]
                        ws_NetworkGroups["C" + str(index_objects)] = json_resp["type"]
                        ws_NetworkGroups["D" + str(index_objects)] = json_resp["description"]
                        ws_NetworkGroups["E" + str(index_objects)] = json_resp["overridable"]
                        ws_NetworkGroups["F" + str(index_objects)] = json_resp["metadata"]["readOnly"]["reason"]

                        ws_NetworkGroups["I" + str(index_objects)] = objects["id"]
                        ws_NetworkGroups["J" + str(index_objects)] = objects["name"]
                        ws_NetworkGroups["K" + str(index_objects)] = objects["type"]

                        index_objects += 1

                    if index_literals > index_objects:
                        index = index_literals
                    else :
                        index = index_objects

                elif "literals" in json_resp:
                    print ("************** literals **************") 
                    for literals in json_resp["literals"]:
                        # print ("literals type : " + literals["type"])
                        # print ("literals value : " + literals["value"])

                        ws_NetworkGroups["A" + str(index)] = json_resp["id"]
                        ws_NetworkGroups["B" + str(index)] = json_resp["name"]
                        ws_NetworkGroups["C" + str(index)] = json_resp["type"]
                        ws_NetworkGroups["D" + str(index)] = json_resp["description"]
                        ws_NetworkGroups["E" + str(index)] = json_resp["overridable"]
                        ws_NetworkGroups["F" + str(index)] = json_resp["metadata"]["readOnly"]["reason"]

                        ws_NetworkGroups["G" + str(index)] = literals["type"]
                        ws_NetworkGroups["H" + str(index)] = literals["value"]
                       
                        index += 1

                elif "objects" in json_resp:
                    print ("************** objects **************") 
                    for objects in json_resp["objects"]:
                        # print ("objects id : " + objects["id"])
                        # print ("objects name : " + objects["name"])
                        # print ("objects type : " + objects["type"])

                        ws_NetworkGroups["A" + str(index)] = json_resp["id"]
                        ws_NetworkGroups["B" + str(index)] = json_resp["name"]
                        ws_NetworkGroups["C" + str(index)] = json_resp["type"]
                        ws_NetworkGroups["D" + str(index)] = json_resp["description"]
                        ws_NetworkGroups["E" + str(index)] = json_resp["overridable"]
                        ws_NetworkGroups["F" + str(index)] = json_resp["metadata"]["readOnly"]["reason"]

                        ws_NetworkGroups["I" + str(index)] = objects["id"]
                        ws_NetworkGroups["J" + str(index)] = objects["name"]
                        ws_NetworkGroups["K" + str(index)] = objects["type"]

                        index += 1
                else :
                    pass

        paging += 1

    print ("\n")
    print ("NetworkGroups Total = " + str(count))
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    print("The Desktop path is: " + desktop)
    wb.save(desktop + "/NetworkGroups.xlsx")


## Function API Get Ports ##
def get_ports():

    offset = 0
    limit = 1000
    index = 2
    paging = 1
    
    ## Create workbook, sheet in Excel ##
    wb = Workbook()
    ws_Ports = wb.active
    ws_Ports.title = "PortObjects"

    ## Header NetworkGroups in Excel ##
    ws_Ports["A1"] = "ID_Ports"
    ws_Ports["A1"].font = Font(bold=True)
    ws_Ports["B1"] = "Name"
    ws_Ports["B1"].font = Font(bold=True)
    ws_Ports["C1"] = "Type"
    ws_Ports["C1"].font = Font(bold=True)
    ws_Ports["D1"] = "Protocal"
    ws_Ports["D1"].font = Font(bold=True)
    ws_Ports["E1"] = "Ports"
    ws_Ports["E1"].font = Font(bold=True)
    ws_Ports["F1"] = "Description"
    ws_Ports["F1"].font = Font(bold=True)
    ws_Ports["G1"] = "Overridable"
    ws_Ports["G1"].font = Font(bold=True)
    ws_Ports["H1"] = "Status"
    ws_Ports["H1"].font = Font(bold=True)

    api_path = "/api/fmc_config/v1/domain/" + g_domain_uuid + "/object/ports?offset=" + str(offset) + "&limit=" +str(limit)    # param
    url = g_server + api_path

    json_resp = get_url(url)

    count = json_resp["paging"]["count"]
    print ("************ Pages total : " + str(json_resp["paging"]["pages"]) + " ************")

    for offset_num in range (offset,count,limit):
        print ("************ paging = " + str(paging) + ", offset = " + str(offset_num) + ", limit = " + str(limit) + " ************")
        api_path = "/api/fmc_config/v1/domain/" + g_domain_uuid + "/object/ports?offset=" + str(offset_num) + "&limit=" +str(limit)    # param
        url = g_server + api_path

        json_resp = get_url(url)

        for data in json_resp["items"]:
            #index = index + 1
            #print (json.dumps(data['id'],sort_keys=True,indent=4, separators=(',', ': ')))
            print ("---------------------------------------------")
            #print ("links : " + data["links"]["self"])
            #print ("index : " + str(index))
            #print ("id : " + data["id"])
            #print ("name : " + data["name"])
            #print ("type : " + data["type"])

            # Get value in networkaddresses
            url = data["links"]["self"]    # param

            json_resp = get_url(url)

            print(json.dumps(json_resp,sort_keys=True,indent=4, separators=(',', ': ')))

            if not "readOnly" in json_resp["metadata"]:
                #print(json.dumps(json_resp,sort_keys=True,indent=4, separators=(',', ': ')))

                if "ProtocolPortObject" in json_resp["type"]:
                    #print ("++++++++++++++++++++++++++ ProtocolPortObject ++++++++++++++++++++++++++")
                    if not "port" in json_resp:
                        # print ("id : " + json_resp["id"])
                        # print ("name : " + json_resp["name"])
                        # print ("type : " + json_resp["type"])
                        # print ("protocol : " + json_resp["protocol"])
                        # print ("description : " + json_resp["description"])
                        # print ("overridable : " + str(json_resp["overridable"]))

                        ws_Ports["A" + str(index)] = json_resp["id"]
                        ws_Ports["B" + str(index)] = json_resp["name"]
                        ws_Ports["C" + str(index)] = json_resp["type"]
                        ws_Ports["D" + str(index)] = json_resp["protocol"]
                        #ws_Ports["E" + str(index)] = json_resp["port"]
                        ws_Ports["F" + str(index)] = json_resp["description"]
                        ws_Ports["G" + str(index)] = str(json_resp["overridable"])

                        index += 1

                    else :
                        # print ("id : " + json_resp["id"])
                        # print ("name : " + json_resp["name"])
                        # print ("type : " + json_resp["type"])
                        # print ("protocol : " + json_resp["protocol"])
                        # print ("port : " + json_resp["port"])
                        # print ("description : " + json_resp["description"])
                        # print ("overridable : " + str(json_resp["overridable"]))

                        ws_Ports["A" + str(index)] = json_resp["id"]
                        ws_Ports["B" + str(index)] = json_resp["name"]
                        ws_Ports["C" + str(index)] = json_resp["type"]
                        ws_Ports["D" + str(index)] = json_resp["protocol"]
                        ws_Ports["E" + str(index)] = json_resp["port"]
                        ws_Ports["F" + str(index)] = json_resp["description"]
                        ws_Ports["G" + str(index)] = str(json_resp["overridable"])

                        index += 1

                #elif "ICMPV4Object" in json_resp["type"]:
                #    print ("++++++++++++++++++++++++++ ICMPV4Object ++++++++++++++++++++++++++")

                else :
                    pass
            
            else :

                #print(json.dumps(json_resp,sort_keys=True,indent=4, separators=(',', ': ')))

                if "ProtocolPortObject" in json_resp["type"]:
                    #print ("++++++++++++++++++++++++++ ProtocolPortObject ++++++++++++++++++++++++++")
                    if not "port" in json_resp:
                        # print ("id : " + json_resp["id"])
                        # print ("name : " + json_resp["name"])
                        # print ("type : " + json_resp["type"])
                        # print ("protocol : " + json_resp["protocol"])
                        # print ("description : " + json_resp["description"])
                        # print ("overridable : " + str(json_resp["overridable"]))
                        # print ("reason : " + json_resp["metadata"]["readOnly"]["reason"])

                        ws_Ports["A" + str(index)] = json_resp["id"]
                        ws_Ports["B" + str(index)] = json_resp["name"]
                        ws_Ports["C" + str(index)] = json_resp["type"]
                        ws_Ports["D" + str(index)] = json_resp["protocol"]
                        #ws_Ports["E" + str(index)] = json_resp["port"]
                        ws_Ports["F" + str(index)] = json_resp["description"]
                        ws_Ports["G" + str(index)] = str(json_resp["overridable"])
                        ws_Ports["H" + str(index)] = json_resp["metadata"]["readOnly"]["reason"]

                        index += 1

                    else :
                        # print ("id : " + json_resp["id"])
                        # print ("name : " + json_resp["name"])
                        # print ("type : " + json_resp["type"])
                        # print ("protocol : " + json_resp["protocol"])
                        # print ("port : " + json_resp["port"])
                        # print ("description : " + json_resp["description"])
                        # print ("overridable : " + str(json_resp["overridable"]))
                        # print ("reason : " + json_resp["metadata"]["readOnly"]["reason"])

                        ws_Ports["A" + str(index)] = json_resp["id"]
                        ws_Ports["B" + str(index)] = json_resp["name"]
                        ws_Ports["C" + str(index)] = json_resp["type"]
                        ws_Ports["D" + str(index)] = json_resp["protocol"]
                        ws_Ports["E" + str(index)] = json_resp["port"]
                        ws_Ports["F" + str(index)] = json_resp["description"]
                        ws_Ports["G" + str(index)] = str(json_resp["overridable"])
                        ws_Ports["H" + str(index)] = json_resp["metadata"]["readOnly"]["reason"]

                        index += 1

                #elif "ICMPV4Object" in json_resp["type"]:
                #    print ("++++++++++++++++++++++++++ ICMPV4Object ++++++++++++++++++++++++++")

                else :
                    pass

        paging += 1

    print ("\n")
    print ("PortObjects (Only ProtocolPortObject) Total = " + str(count))
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    print("The Desktop path is: " + desktop)
    wb.save(desktop + "/PortObjects.xlsx")

def get_portgroups():

    offset = 0
    limit = 1000
    index = 2
    paging = 1

    ## Create workbook, sheet in Excel ##
    wb = Workbook()
    ws_PortGroups = wb.active
    ws_PortGroups.title = "PortGroups"

    ## Header NetworkGroups in Excel ##
    ws_PortGroups["A1"] = "ID_Groups"
    ws_PortGroups["A1"].font = Font(bold=True)
    ws_PortGroups["B1"] = "Name"
    ws_PortGroups["B1"].font = Font(bold=True)
    ws_PortGroups["C1"] = "Type"
    ws_PortGroups["C1"].font = Font(bold=True)
    ws_PortGroups["D1"] = "Description"
    ws_PortGroups["D1"].font = Font(bold=True)
    ws_PortGroups["E1"] = "Overridable"
    ws_PortGroups["E1"].font = Font(bold=True)
    ws_PortGroups["F1"] = "Object_ID"
    ws_PortGroups["F1"].font = Font(bold=True)
    ws_PortGroups["G1"] = "Object_Name"
    ws_PortGroups["G1"].font = Font(bold=True)
    ws_PortGroups["H1"] = "Object_Type"
    ws_PortGroups["H1"].font = Font(bold=True)


    api_path = "/api/fmc_config/v1/domain/" + g_domain_uuid + "/object/portobjectgroups?offset=" + str(offset) + "&limit=" +str(limit)    # param
    url = g_server + api_path

    json_resp = get_url(url)

    count = json_resp["paging"]["count"]
    print ("************ Pages total : " + str(json_resp["paging"]["pages"]) + " ************")

    for offset_num in range (offset,count,limit):
        print ("************ paging = " + str(paging) + ", offset = " + str(offset_num) + ", limit = " + str(limit) + " ************")
        api_path = "/api/fmc_config/v1/domain/" + g_domain_uuid + "/object/portobjectgroups?offset=" + str(offset_num) + "&limit=" +str(limit)    # param
        url = g_server + api_path

        json_resp = get_url(url)

        for data in json_resp["items"]:
            #index = index + 1
            #print (json.dumps(data['id'],sort_keys=True,indent=4, separators=(',', ': ')))
            print ("---------------------------------------------")
            #print ("links : " + data["links"]["self"])
            #print ("index : " + str(index))
            #print ("id : " + data["id"])
            #print ("name : " + data["name"])
            #print ("type : " + data["type"])

            url = data["links"]["self"]    # param

            json_resp = get_url(url)

            print(json.dumps(json_resp,sort_keys=True,indent=4, separators=(',', ': ')))

            for objects in json_resp["objects"]:
                # print ("********************************************")
                # print ("id : " + objects["id"])
                # print ("name : " + objects["name"])
                # print ("type : " + objects["type"])

                ws_PortGroups["A" + str(index)] = json_resp["id"]
                ws_PortGroups["B" + str(index)] = json_resp["name"]
                ws_PortGroups["C" + str(index)] = json_resp["type"]
                ws_PortGroups["D" + str(index)] = json_resp["description"]
                ws_PortGroups["E" + str(index)] = str(json_resp["overridable"])
                ws_PortGroups["F" + str(index)] = objects["id"]
                ws_PortGroups["G" + str(index)] = objects["name"]
                ws_PortGroups["H" + str(index)] = objects["type"]

                index += 1


    print ("\n")
    print ("PortGroups Total = " + str(count))
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    print("The Desktop path is: " + desktop)
    wb.save(desktop + "/PortGroups.xlsx")


# def check_paging(url,headers):

#     try:
#         r = requests.get(url, headers=headers, verify=False)
#         status_code = r.status_code
#         resp = r.text
#         if (status_code == 200):
#             print("GET successful. Response data --> ")
#             json_resp = json.loads(resp)

#             #print(json.dumps(json_resp,sort_keys=True,indent=4, separators=(',', ': ')))
#             print(json.dumps(json_resp["paging"],sort_keys=True,indent=4, separators=(',', ': ')))

#             count = json_resp["paging"]["count"]

#         else:
#             r.raise_for_status()
#             print("Error occurred in GET --> "+resp)

#     except requests.exceptions.HTTPError as err :
#         print ("Error in connection --> "+str(err))
#     finally:
#         if r : 
#             r.close()
#             return count

################ Main Function ################



# if __name__ == '__main__':
#     server = "https://xxx.xxx.xxx.xxx"
#     username = "username"
#     password = "password"

#     connect(server, username, password)

#     get_networkaddresses()
#     get_networkgroups()
#     get_ports()
#     get_portgroups()

